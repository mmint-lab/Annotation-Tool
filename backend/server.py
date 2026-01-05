from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, HTMLResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorGridFSBucket
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta, timezone
import jwt
import bcrypt
import csv
import io
import json
import mammoth
import re
from collections import defaultdict
from urllib.parse import urlparse
from bson import ObjectId

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-this-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Allowed file extensions for resource uploads
ALLOWED_RESOURCE_EXT = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'txt', 'png', 'jpg', 'jpeg', 'gif', 'svg'}

# Create the main app without a prefix
app = FastAPI(title="Social Determinants of Health Annotation Tool")

# Create a router with the /api prefix
# GridFS bucket for resources
fs_bucket = AsyncIOMotorGridFSBucket(db)

api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# User Roles
class UserRole:
    ADMIN = "admin"
    ANNOTATOR = "annotator"

# Social Determinants of Health Domains and Structured Tags
SDOH_DOMAINS = [
    "Economic Stability",
    "Education Access and Quality", 
    "Health Care Access and Quality",
    "Neighborhood and Built Environment",
    "Social and Community Context"
]

# Structured tag definitions (placeholder)
SDOH_TAG_STRUCTURE = {
    "Economic Stability": {
        "Employment": [
            "Employed", "Disabled",
            "Retired", "Homemaker", "Student", "Harmful Workplace"
        ],
        "Food Insecurity": [
            "Low or Very Low Food Security",
            "Physical Access Barrier", "Food Assistance Program"
        ],
        "Housing Instability": [
            "Multiple Moves",
            "Eviction or Foreclosure", "Unhoused",
            "Housing Assistance Program"
        ],
        "Poverty": [
            "Low Socioeconomic Status",
            "Social Assistance Program"
        ]
    },
    "Education Access and Quality": {
        "Early Childhood Development and Education": [
            "School Readiness Concern", "Developmental Delay",
            "Reading Impairment", "Math Impairment", "Other Learning Disability",
            "Early Intervention Services", "Learning Environment Concern"
        ],
        "Highest Level of Education": [
            "Some High School", "High School Diploma or GED",
            "Some College or Associate Degree", "Bachelor's Degree",
            "Graduate or Professional Degree"
        ],
        "Language and Literacy": [
            "Language Barrier"
        ]
    },
    "Health Care Access and Quality": {
        "Access to Health Services": [
            "Insurance Coverage Limitations",
            "Limited Access to Specialist Health Services"
        ],
        "Access to Primary Care": [
            "No Primary Care Provider", "Extended Gaps in Care",
            "Geographic Barriers to Care", "Limited Appointment Availability"
        ],
        "Health Literacy": [
            "Difficulty Understanding Medical Information",
            "Digital Health Literacy Gaps"
        ]
    },
    "Neighborhood and Built Environment": {
        "Access to Healthy Foods": [
            "Limited Healthy Food Options",
            "Increased Distance to Food Sources"
        ],
        "Crime and Violence": [
            "Perceived Neighborhood Safety", "Reported Criminal Activity"
        ],
        "Environmental Conditions": [
            "Air Quality Concerns", "Water Quality Issues",
            "Proximity to Environmental Hazards"
        ],
        "Quality of Housing": [
            "Structural Deficiencies", "Pest Infestation or Mold",
            "Overcrowding or Unsafe Occupancy"
        ]
    },
    "Social and Community Context": {
        "Civic Participation": [
            "Voting and Political Engagement", "Community or Religious Involvement",
            "Volunteering or Service Activities"
        ],
        "Incarceration": [
            "Current Incarceration", "History of Incarceration",
            "Community Supervision"
        ],
        "Social Cohesion": [
            "Supportive Relationships", "Trust and Belonging in Community"
        ],
        "Experiences of Discrimination or Exclusion": [
            "Race", "Ethnicity", "Gender", "Sexual Orientation", "Other Identity"
        ]
    }
}

# ========================
# Models (trimmed for brevity) - ensure these exist in your real file
# ========================
class Token(BaseModel):
    access_token: str
    token_type: str = "bearer"

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    full_name: str = ""
    role: str = UserRole.ANNOTATOR

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str = ""
    role: str = UserRole.ANNOTATOR

class UserLogin(BaseModel):
    email: str  # Can be email or username (full_name)
    password: str

class AnnotationTag(BaseModel):
    domain: str
    category: Optional[str] = None
    tag: str
    valence: Optional[str] = "positive"
    confidence: Optional[int] = 3  # 1-5 scale, per-tag confidence (1=least confident, 5=most confident)

class AnnotationCreate(BaseModel):
    sentence_id: str
    tags: List[AnnotationTag] = []
    notes: Optional[str] = ""
    skipped: bool = False
    confidence: Optional[int] = None  # 1-5 scale (deprecated, use per-tag confidence)
    duration_ms: Optional[int] = None  # time spent selecting tags for this sentence in ms

class Annotation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sentence_id: str
    user_id: str
    tags: List[AnnotationTag] = []
    notes: Optional[str] = ""
    skipped: bool = False
    confidence: Optional[int] = None  # 0-5 scale
    duration_ms: Optional[int] = None  # time spent selecting tags for this sentence in ms
    created_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())

class UserActivity(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    document_id: Optional[str] = None
    sentence_id: Optional[str] = None
    action_type: str  # page_navigation, tag_click, sentence_transition
    timestamp: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    metadata: Optional[Dict[str, Any]] = {}  # Additional context like tag clicked, direction moved, etc.

class UserActivityCreate(BaseModel):
    document_id: Optional[str] = None
    sentence_id: Optional[str] = None
    action_type: str
    metadata: Optional[Dict[str, Any]] = {}


# ========================
# Auth helpers
# ========================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

async def get_default_project_name() -> str:
    """Get the default project name from settings (db.settings) or environment"""
    s = await db.settings.find_one({"key": "default_project_name"}, {"_id": 0})
    if s and s.get('value'):
        return s['value']
    return os.environ.get('DEFAULT_PROJECT_NAME', 'Default Project')

async def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        return user_id
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(user_id: str = Depends(verify_token)):
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return User(**user)

async def get_admin_user(current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def get_current_user_optional(credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False))):
    """Optional authentication - returns None if no token provided"""
    if not credentials:
        return None
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            return None
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            return None
        if isinstance(user.get('created_at'), datetime):
            user['created_at'] = user['created_at'].isoformat()
        return User(**user)
    except jwt.PyJWTError:
        return None

# ========================
# Token creation
# ========================

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# ========================
# Basic endpoints
# ========================
@api_router.get("/")
async def root():
    return {"status": "ok"}

@api_router.get("/domains")
async def get_domains():
    """Get SDOH domains"""
    return SDOH_DOMAINS

@api_router.get("/tag-structure")
async def get_tag_structure():
    return SDOH_TAG_STRUCTURE

# ========================
# Authentication endpoints
# ========================
@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate):
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    user = User(email=user_data.email, full_name=user_data.full_name, role=user_data.role)
    user_dict = user.dict()
    user_dict["password"] = hash_password(user_data.password)
    await db.users.insert_one(user_dict)
    return user

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    # Try to find user by email first, then by full_name (username)
    user = await db.users.find_one({"email": user_data.email})
    if not user:
        # Try finding by full_name as username
        user = await db.users.find_one({"full_name": user_data.email})
    
    if not user or not verify_password(user_data.password, user.get("password", "")):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    access_token = create_access_token(data={"sub": user["id"]})
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

class ProfileUpdateRequest(BaseModel):
    full_name: str

@api_router.put("/auth/me/profile")
async def update_profile(data: ProfileUpdateRequest, current_user: User = Depends(get_current_user)):
    await db.users.update_one({"id": current_user.id}, {"$set": {"full_name": data.full_name}})
    user = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    return User(**user)

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_password(data: ChangePasswordRequest, current_user: User = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user.id})
    if not user or not verify_password(data.current_password, user.get("password", "")):
        raise HTTPException(status_code=400, detail="Current password incorrect")
    new_hash = hash_password(data.new_password)
    await db.users.update_one({"id": current_user.id}, {"$set": {"password": new_hash}})
    return {"status": "ok"}

# ========================
# Document endpoints
# ========================
@api_router.get("/documents")
async def get_documents(current_user: User = Depends(get_current_user)):
    documents = await db.documents.find({}, {"_id": 0}).to_list(1000)
    return documents

@api_router.post("/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    project_name: Optional[str] = None,
    description: Optional[str] = None,
    current_user: User = Depends(get_admin_user)
):
    # Read CSV
    content = await file.read()
    csv_content = content.decode('utf-8', errors='ignore')
    reader = csv.DictReader(io.StringIO(csv_content))

    def pick_text(row: Dict[str, Any]) -> str:
        # First check for specific medical record columns
        text_parts = []
        if 'HISTORY OF PRESENT ILLNESS' in row and isinstance(row['HISTORY OF PRESENT ILLNESS'], str) and row['HISTORY OF PRESENT ILLNESS'].strip():
            text_parts.append(row['HISTORY OF PRESENT ILLNESS'].strip())
        if 'SOCIAL HISTORY' in row and isinstance(row['SOCIAL HISTORY'], str) and row['SOCIAL HISTORY'].strip():
            text_parts.append(row['SOCIAL HISTORY'].strip())
        
        if text_parts:
            return ' '.join(text_parts)
        
        # Try common text columns as fallback
        for key in ['discharge_summary','text','note','notes','summary','content','sentence']:
            if key in row and isinstance(row[key], str) and row[key].strip():
                return row[key].strip()
        
        # fallback: join fields that look like text
        acc = []
        for k, v in row.items():
            if isinstance(v, str) and len(v.strip()) > 20:
                acc.append(v.strip())
        return ' '.join(acc)

    def pick_subject(row: Dict[str, Any], idx: int) -> str:
        # Prioritize note_id for medical records
        if 'note_id' in row and str(row['note_id']).strip():
            return str(row['note_id']).strip()
        # Then try other common ID columns
        for key in ['subject_id','patient_id','encounter_id','index','id']:
            if key in row and str(row[key]).strip():
                return str(row[key]).strip()
        return str(idx)

    def split_sentences(text: str) -> List[str]:
        # conservative split keeping content
        parts = re.split(r'(?:[.!?]+\s+)', text)
        # If re.split kept separators out, it's okay; clean empties
        return [p.strip() for p in parts if p and p.strip()]

    sentences = []
    row_idx = 0
    for row in reader:
        row_idx += 1
        text = pick_text(row)
        if not text:
            continue
        chunks = split_sentences(text)
        subj = pick_subject(row, row_idx)
        for si, chunk in enumerate(chunks):
            sentences.append({
                'id': str(uuid.uuid4()),
                'text': chunk,
                'subject_id': subj,
                'row_index': row_idx,
                'sentence_index': si,
                'document_id': '',
                'created_at': datetime.utcnow().isoformat()
            })

    document_id = str(uuid.uuid4())
    document = {
        'id': document_id,
        'filename': file.filename,
        'project_name': await get_default_project_name(),  # use default project from settings
        'description': description,
        'total_sentences': len(sentences),
        'uploaded_by': current_user.id,
        'created_at': datetime.utcnow().isoformat(),
        'last_modified_by': {}  # Will store {user_id: timestamp} for per-user tracking
    }

    for s in sentences:
        s['document_id'] = document_id

    await db.documents.insert_one(document)
    if sentences:
        await db.sentences.insert_many(sentences)

    return {k: document[k] for k in ['id','filename','project_name','description','total_sentences','uploaded_by','created_at']}

@api_router.get("/documents/{document_id}/sentences")
async def get_document_sentences(
    document_id: str,
    current_user: User = Depends(get_current_user)
):
    sentences = await db.sentences.find({"document_id": document_id}, {"_id": 0}).to_list(1000)
    for sentence in sentences:
        annotations = await db.annotations.find({"sentence_id": sentence["id"]}, {"_id": 0}).to_list(100)
        sentence["annotations"] = annotations
    return sentences

# ========================
# Annotation endpoints
# ========================
@api_router.post("/annotations", response_model=Annotation)
async def create_annotation(annotation_data: AnnotationCreate, current_user: User = Depends(get_current_user)):
    # Enforce per-user, per-sentence exclusivity: either 'skip' or 'tags', not both
    if annotation_data.skipped:
        # Remove any existing annotations (tagged or skipped) for this user on this sentence
        await db.annotations.delete_many({
            "sentence_id": annotation_data.sentence_id,
            "user_id": current_user.id
        })
        annotation = Annotation(
            sentence_id=annotation_data.sentence_id,
            user_id=current_user.id,
            tags=[],
            notes=annotation_data.notes,
            skipped=True,
            confidence=annotation_data.confidence,
            duration_ms=annotation_data.duration_ms
        )
        await db.annotations.insert_one(annotation.dict())
    else:
        # Remove any existing 'skip' annotation by this user for the sentence, keep other tagged annotations if desired
        await db.annotations.delete_many({
            "sentence_id": annotation_data.sentence_id,
            "user_id": current_user.id,
            "skipped": True
        })
        annotation = Annotation(
            sentence_id=annotation_data.sentence_id,
            user_id=current_user.id,
            tags=annotation_data.tags,
            notes=annotation_data.notes,
            skipped=False,
            confidence=annotation_data.confidence,
            duration_ms=annotation_data.duration_ms
        )
        await db.annotations.insert_one(annotation.dict())
    
    # Update document's last_modified_by timestamp for this user
    sentence = await db.sentences.find_one({"id": annotation_data.sentence_id})
    if sentence:
        document_id = sentence.get("document_id")
        if document_id:
            await db.documents.update_one(
                {"id": document_id},
                {"$set": {f"last_modified_by.{current_user.id}": datetime.utcnow().isoformat()}}
            )
    
    return annotation

@api_router.delete("/annotations/{annotation_id}")
async def delete_annotation(annotation_id: str, current_user: User = Depends(get_current_user)):
    # Allow delete if admin or owner
    ann = await db.annotations.find_one({"id": annotation_id})
    if not ann:
        raise HTTPException(status_code=404, detail="Annotation not found")
    if current_user.role != UserRole.ADMIN and ann.get("user_id") != current_user.id:
        raise HTTPException(status_code=403, detail="Not allowed")
    await db.annotations.delete_one({"id": annotation_id})
    return {"deleted": 1}

class RemoveTagPayload(BaseModel):
    domain: str
    category: Optional[str] = None
    tag: str

@api_router.post("/annotations/{annotation_id}/remove-tag")
async def remove_tag_from_annotation(annotation_id: str, payload: RemoveTagPayload, current_user: User = Depends(get_current_user)):
    ann = await db.annotations.find_one({"id": annotation_id})
    if not ann:
        raise HTTPException(status_code=404, detail="Annotation not found")
    if current_user.role != UserRole.ADMIN and ann.get("user_id") != current_user.id:
        raise HTTPException(status_code=403, detail="Not allowed")
    tags = ann.get("tags", [])
    def norm(x):
        return (x or "").strip().lower()
    pdomain, pcat, ptag = norm(payload.domain), norm(payload.category), norm(payload.tag)
    def tag_matches(t):
        if isinstance(t, dict):
            return norm(t.get("domain")) == pdomain and norm(t.get("category")) == pcat and norm(t.get("tag")) == ptag
        else:
            return norm(getattr(t, "domain", None)) == pdomain and norm(getattr(t, "category", None)) == pcat and norm(getattr(t, "tag", None)) == ptag
    new_tags = [t for t in tags if not tag_matches(t)]
    if not new_tags and not ann.get("skipped", False):
        await db.annotations.delete_one({"id": annotation_id})
        return {"deleted": 1}
    await db.annotations.update_one({"id": annotation_id}, {"$set": {"tags": new_tags}})
    updated = await db.annotations.find_one({"id": annotation_id}, {"_id": 0})
    return updated

@api_router.get("/annotations/sentence/{sentence_id}")
async def get_sentence_annotations(sentence_id: str, current_user: User = Depends(get_current_user)):
    annotations = await db.annotations.find({"sentence_id": sentence_id}, {"_id": 0}).to_list(100)
    return annotations

@api_router.post("/annotations/bulk-delete")
async def bulk_delete_annotations(payload: Dict[str, List[str]], current_user: User = Depends(get_current_user)):
    ids = payload.get("annotation_ids", [])
    if ids:
        await db.annotations.delete_many({"id": {"$in": ids}})
    return {"deleted": len(ids)}

@api_router.delete("/annotations/document/{document_id}/clear-all")
async def clear_all_document_annotations(document_id: str, current_user: User = Depends(get_current_user)):
    """Clear all annotations for a document (only current user's annotations unless admin)"""
    # Get all sentence IDs for this document
    sentence_ids = await db.sentences.distinct("id", {"document_id": document_id})
    if not sentence_ids:
        raise HTTPException(status_code=404, detail="Document not found or has no sentences")
    
    # Build query - admins can delete all, regular users only their own
    query = {"sentence_id": {"$in": sentence_ids}}
    if current_user.role != UserRole.ADMIN:
        query["user_id"] = current_user.id
    
    result = await db.annotations.delete_many(query)
    return {"deleted": result.deleted_count}


@api_router.get("/annotations/active-docs")
async def get_active_docs(scope: str = Query("me"), current_user: User = Depends(get_current_user)):
    """Get documents with annotation progress for the active documents panel"""
    # Get all documents
    docs = await db.documents.find({}, {"_id": 0}).to_list(10000)
    result = []
    
    for doc in docs:
        # Get total sentences for this document
        total_sentences = await db.sentences.count_documents({"document_id": doc["id"]})
        if total_sentences == 0:
            continue
        
        # Get sentence IDs for this document
        sentence_ids = await db.sentences.distinct("id", {"document_id": doc["id"]})
        
        # Filter annotations by scope
        ann_filter = {"sentence_id": {"$in": sentence_ids}}
        if scope == "me":
            ann_filter["user_id"] = current_user.id
        
        # Count annotated sentences
        annotated_sentence_ids = await db.annotations.distinct("sentence_id", ann_filter)
        annotated_count = len(annotated_sentence_ids)
        
        # Calculate progress
        progress = annotated_count / total_sentences if total_sentences > 0 else 0
        
        # Get last annotation index for resume functionality
        last_annotation = await db.annotations.find(ann_filter, {"_id": 0}).sort("created_at", -1).limit(1).to_list(1)
        last_annotation_index = None
        if last_annotation:
            last_sentence = await db.sentences.find_one({"id": last_annotation[0]["sentence_id"]}, {"_id": 0})
            if last_sentence:
                last_annotation_index = last_sentence.get("index", 0)
        
        # Only include documents with at least one annotation or if admin viewing team
        if annotated_count > 0 or (scope == "team" and current_user.role == UserRole.ADMIN):
            result.append({
                "document_id": doc["id"],
                "filename": doc.get("filename", ""),
                "total_sentences": total_sentences,
                "annotated_count": annotated_count,
                "progress": progress,
                "last_annotation_index": last_annotation_index
            })
    
    # Sort by most recently annotated
    result.sort(key=lambda x: x.get("last_annotation_index") or 0, reverse=True)
    return result


# ========================
# User Activity Tracking
# ========================
@api_router.post("/activities")
async def log_activity(activity_data: UserActivityCreate, current_user: User = Depends(get_current_user)):
    """Log user activity for timestamp tracking"""
    activity = UserActivity(
        user_id=current_user.id,
        document_id=activity_data.document_id,
        sentence_id=activity_data.sentence_id,
        action_type=activity_data.action_type,
        metadata=activity_data.metadata or {}
    )
    await db.user_activities.insert_one(activity.dict())
    return {"status": "logged"}

@api_router.get("/admin/download/activity-log")
async def download_activity_log(
    document_id: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    current_user: User = Depends(get_admin_user)
):
    """Download activity log as CSV (admin only)"""
    query: Dict[str, Any] = {}
    if document_id:
        query['document_id'] = document_id
    if user_id:
        query['user_id'] = user_id
    
    activities = await db.user_activities.find(query, {"_id": 0}).sort("timestamp", -1).to_list(100000)
    
    # Get user display names
    user_ids = list(set(a.get('user_id') for a in activities if a.get('user_id')))
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0}).to_list(1000)
    user_display = {u['id']: (u.get('full_name') or u.get('email') or u['id']) for u in users}
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["timestamp", "user_id", "user_name", "document_id", "sentence_id", "action_type", "metadata"])
    
    for a in activities:
        writer.writerow([
            a.get('timestamp', ''),
            a.get('user_id', ''),
            user_display.get(a.get('user_id', ''), ''),
            a.get('document_id', ''),
            a.get('sentence_id', ''),
            a.get('action_type', ''),
            json.dumps(a.get('metadata', {}))
        ])
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=activity_log.csv"}
    )


# ========================
# Admin endpoints
# ========================

# Admin User Management Models
class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

class BulkDeleteRequest(BaseModel):
    ids: List[str]

# Admin Users Endpoints
@api_router.get("/admin/users")
async def get_all_users(current_user: User = Depends(get_admin_user)):
    """Get all users (admin only) - returns users without passwords"""
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    # Ensure required fields exist with defaults
    for user in users:
        if "is_active" not in user:
            user["is_active"] = True
        if "role" not in user:
            user["role"] = UserRole.ANNOTATOR  # Default role for legacy users
    return users

@api_router.post("/admin/users", response_model=User)
async def create_user(user_data: UserCreate, current_user: User = Depends(get_admin_user)):
    """Create a new user (admin only)"""
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        email=user_data.email, 
        full_name=user_data.full_name, 
        role=user_data.role
    )
    user_dict = user.dict()
    user_dict["password"] = hash_password(user_data.password)
    user_dict["is_active"] = True  # New users are active by default
    
    await db.users.insert_one(user_dict)
    return user

@api_router.put("/admin/users/{user_id}")
async def update_user(user_id: str, user_update: UserUpdate, current_user: User = Depends(get_admin_user)):
    """Update user (admin only) - can update is_active, role, full_name"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {}
    if user_update.full_name is not None:
        update_data["full_name"] = user_update.full_name
    if user_update.role is not None:
        if user_update.role not in [UserRole.ADMIN, UserRole.ANNOTATOR]:
            raise HTTPException(status_code=400, detail="Invalid role")
        update_data["role"] = user_update.role
    if user_update.is_active is not None:
        update_data["is_active"] = user_update.is_active
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    # Ensure is_active field exists
    if "is_active" not in updated_user:
        updated_user["is_active"] = True
    return updated_user

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_admin_user)):
    """Delete user (admin only) - prevents self-delete"""
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Delete user's annotations first
    annotations_deleted = await db.annotations.delete_many({"user_id": user_id})
    
    # Delete the user
    await db.users.delete_one({"id": user_id})
    
    return {
        "message": "User deleted successfully",
        "user_name": user.get("full_name", user.get("email", "")),
        "annotations_deleted": annotations_deleted.deleted_count
    }

@api_router.post("/admin/users/bulk-delete")
async def bulk_delete_users(request: BulkDeleteRequest, current_user: User = Depends(get_admin_user)):
    """Bulk delete users (admin only) - skips current admin"""
    user_ids = [uid for uid in request.ids if uid != current_user.id]  # Skip current admin
    
    if not user_ids:
        return {"deleted": 0, "skipped": len(request.ids)}
    
    # Get user info before deletion
    users_to_delete = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0}).to_list(1000)
    
    # Delete annotations for these users
    annotations_deleted = await db.annotations.delete_many({"user_id": {"$in": user_ids}})
    
    # Delete the users
    result = await db.users.delete_many({"id": {"$in": user_ids}})
    
    return {
        "deleted": result.deleted_count,
        "skipped": len(request.ids) - len(user_ids),
        "annotations_deleted": annotations_deleted.deleted_count,
        "deleted_users": [{"id": u["id"], "name": u.get("full_name", u.get("email", ""))} for u in users_to_delete]
    }

# Admin Documents Endpoints
@api_router.delete("/admin/documents/{document_id}")
async def delete_document(document_id: str, current_user: User = Depends(get_admin_user)):
    """Delete document and cascade to sentences/annotations (admin only)"""
    document = await db.documents.find_one({"id": document_id})
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Get sentence IDs for this document
    sentence_ids = await db.sentences.distinct("id", {"document_id": document_id})
    
    # Delete annotations for these sentences
    annotations_deleted = 0
    if sentence_ids:
        result = await db.annotations.delete_many({"sentence_id": {"$in": sentence_ids}})
        annotations_deleted = result.deleted_count
    
    # Delete sentences
    sentences_deleted = await db.sentences.delete_many({"document_id": document_id})
    
    # Delete the document
    await db.documents.delete_one({"id": document_id})
    
    return {
        "message": "Document deleted successfully",
        "document_name": document.get("filename", ""),
        "sentences_deleted": sentences_deleted.deleted_count,
        "annotations_deleted": annotations_deleted
    }


@api_router.post("/admin/documents/{document_id}/assign-users")
async def assign_users_to_document(document_id: str, request: Dict[str, List[str]], current_user: User = Depends(get_admin_user)):
    """Assign users to a document (admin only)"""
    user_ids = request.get('user_ids', [])
    
    document = await db.documents.find_one({"id": document_id})
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Validate that all user_ids exist
    if user_ids:
        existing_users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1}).to_list(1000)
        existing_user_ids = [u['id'] for u in existing_users]
        invalid_ids = [uid for uid in user_ids if uid not in existing_user_ids]
        if invalid_ids:
            raise HTTPException(status_code=400, detail=f"Invalid user IDs: {invalid_ids}")
    
    # Update document with assigned users
    await db.documents.update_one(
        {"id": document_id},
        {"$set": {"assigned_users": user_ids}}
    )
    
    return {"message": "Users assigned successfully", "assigned_users": user_ids}

@api_router.post("/admin/documents/bulk-delete")
async def bulk_delete_documents(request: Dict[str, List[str]], current_user: User = Depends(get_admin_user)):
    """Bulk delete documents and cascade to sentences/annotations (admin only)"""
    document_ids = request.get('ids') or request.get('document_ids') or []
    
    if not document_ids:
        return {"deleted": 0}
    
    # Get document info before deletion
    documents_to_delete = await db.documents.find({"id": {"$in": document_ids}}, {"_id": 0}).to_list(1000)
    
    # Get all sentence IDs for these documents
    sentence_ids = await db.sentences.distinct("id", {"document_id": {"$in": document_ids}})
    
    # Delete annotations for these sentences
    annotations_deleted = 0
    if sentence_ids:
        result = await db.annotations.delete_many({"sentence_id": {"$in": sentence_ids}})
        annotations_deleted = result.deleted_count
    
    # Delete sentences
    sentences_result = await db.sentences.delete_many({"document_id": {"$in": document_ids}})
    
    # Delete the documents
    documents_result = await db.documents.delete_many({"id": {"$in": document_ids}})
    
    return {
        "deleted": documents_result.deleted_count,
        "sentences_deleted": sentences_result.deleted_count,
        "annotations_deleted": annotations_deleted,
        "deleted_documents": [{"id": d["id"], "name": d.get("filename", "")} for d in documents_to_delete]
    }

@api_router.get("/documents/{document_id}/annotations")
async def get_document_annotations(document_id: str, current_user: User = Depends(get_admin_user)):
    # manager view: list all annotations for a doc
    sentence_ids = await db.sentences.distinct("id", {"document_id": document_id})
    anns = []
    if sentence_ids:
        items = await db.annotations.find({"sentence_id": {"$in": sentence_ids}}, {"_id": 0}).to_list(5000)
        # enrich with sentence text and subject/index for filtering
        text_map = {}
        subj_map = {}
        idx_map = {}
        async for s in db.sentences.find({"document_id": document_id}, {"_id": 0, "id": 1, "text": 1, "subject_id": 1, "index": 1}):
            text_map[s["id"]] = s.get("text")
            subj_map[s["id"]] = s.get("subject_id")
            idx_map[s["id"]] = s.get("index")
        for a in items:
            a["sentence_text"] = text_map.get(a["sentence_id"], "")
            a["subject_id"] = subj_map.get(a["sentence_id"]) or None
            a["sentence_index"] = idx_map.get(a["sentence_id"]) or 0
            anns.append(a)
    return anns

@api_router.get("/admin/download/annotated-csv-inline/{document_id}")
async def download_annotated_csv_inline(document_id: str, user_id: Optional[str] = Query(None), current_user: User = Depends(get_admin_user)):
    document = await db.documents.find_one({"id": document_id})
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    sentences = await db.sentences.find({"document_id": document_id}, {"_id": 0}).sort([("row_index",1),("sentence_index",1)]).to_list(100000)
    sentence_map = {s["id"]: s for s in sentences}
    # Build user display map
    ann_user_ids = await db.annotations.distinct("user_id", {"sentence_id": {"$in": list(sentence_map.keys())}})
    users = await db.users.find({"id": {"$in": ann_user_ids}}, {"_id": 0}).to_list(1000)
    user_display = {u['id']: (u.get('full_name') or u.get('email') or u['id']) for u in users}
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["document_id","sentence_id","subject_id","row_index","sentence_index","sentence_text","tag_domain","tag_category","tag","valence","notes","user_id","user_display","is_skipped","confidence","duration_ms"]) 
    # For each sentence, expand multiple tags into multiple rows; skipped produces a single row with is_skipped True
    for sid, s in sentence_map.items():
        q = {"sentence_id": sid}
        if user_id:
            q["user_id"] = user_id
        anns = await db.annotations.find(q, {"_id": 0}).to_list(1000)
        if not anns:
            writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), "","","","","","","", False,"",""])
            continue
        for a in anns:
            confidence = a.get("confidence", "")
            duration_ms = a.get("duration_ms", "")
            if a.get("skipped"):
                ud = user_display.get(a.get("user_id",""), a.get("user_id",""))
                writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), "","","","", a.get("notes",""), a.get("user_id",""), ud, True, confidence, duration_ms])
                continue
            tags = a.get("tags", [])
            if isinstance(tags, list) and tags:
                for t in tags:
                    domain = t.get("domain") if isinstance(t, dict) else getattr(t, "domain", "")
                    category = t.get("category") if isinstance(t, dict) else getattr(t, "category", "")
                    tag = t.get("tag") if isinstance(t, dict) else getattr(t, "tag", "")
                    valence = t.get("valence") if isinstance(t, dict) else getattr(t, "valence", "")
                    ud = user_display.get(a.get("user_id",""), a.get("user_id",""))
                    writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), domain, category, tag, valence, a.get("notes",""), a.get("user_id",""), ud, False, confidence, duration_ms])
            else:
                ud = user_display.get(a.get("user_id",""), a.get("user_id",""))
                writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), "","","","", a.get("notes",""), a.get("user_id",""), ud, False, confidence, duration_ms])
    output.seek(0)
    return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv", headers={"Content-Disposition": f"inline; filename=annotated_inline_{document['filename']}.csv"})

@api_router.get("/admin/download/annotated-csv-split/{document_id}")
async def download_annotated_csv_split(document_id: str, current_user: User = Depends(get_admin_user)):
    # Tidy format: one row per tag (or one row for skipped)
    document = await db.documents.find_one({"id": document_id})
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    sentences = await db.sentences.find({"document_id": document_id}, {"_id": 0}).sort([("row_index",1),("sentence_index",1)]).to_list(10000)
    sentence_map = {s["id"]: s for s in sentences}
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["document_id","sentence_id","subject_id","row_index","sentence_index","sentence_text","is_skipped","tag_domain","tag_category","tag","valence","notes","user_id"]) 
    cursor = db.annotations.find({"sentence_id": {"$in": list(sentence_map.keys())}}, {"_id": 0})
    anns = await cursor.to_list(100000)
    for a in anns:
        s = sentence_map.get(a["sentence_id"], {})
        if a.get("skipped"):
            writer.writerow([document_id, a.get("sentence_id"), s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), True, "","","","", a.get("notes",""), a.get("user_id","")])
        else:
            tags = a.get("tags", [])
            if isinstance(tags, list) and tags:
                for t in tags:
                    # t might be dict or pydantic model dict
                    domain = t.get("domain") if isinstance(t, dict) else getattr(t, "domain", "")
                    category = t.get("category") if isinstance(t, dict) else getattr(t, "category", "")
                    tag = t.get("tag") if isinstance(t, dict) else getattr(t, "tag", "")
                    valence = t.get("valence") if isinstance(t, dict) else getattr(t, "valence", "")
                    writer.writerow([document_id, a.get("sentence_id"), s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), False, domain, category, tag, valence, a.get("notes",""), a.get("user_id","")])
            else:
                writer.writerow([document_id, a.get("sentence_id"), s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), False, "","","","", a.get("notes",""), a.get("user_id","")])
    output.seek(0)
    return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=annotated_split_{document['filename']}.csv"})

@api_router.get("/admin/download/annotated-paragraphs/{document_id}")
async def download_annotated_paragraphs(document_id: str, current_user: User = Depends(get_admin_user)):
    doc = await db.documents.find_one({"id": document_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    # Fetch sentences ordered by original row/sentence order
    sents = await db.sentences.find({"document_id": document_id}, {"_id": 0}).sort([("row_index",1),("sentence_index",1)]).to_list(100000)
    if not sents:
        raise HTTPException(status_code=404, detail="No sentences for document")
    ids = [s['id'] for s in sents]
    # Prefetch all annotations for these sentences
    anns = await db.annotations.find({"sentence_id": {"$in": ids}}, {"_id": 0}).to_list(200000)
    by_sid: Dict[str, List[Dict[str, Any]]] = {}
    ann_user_ids = set()
    for a in anns:
        by_sid.setdefault(a['sentence_id'], []).append(a)
        if a.get('user_id'):
            ann_user_ids.add(a['user_id'])
    # Build user display map (full_name or email)
    users = await db.users.find({"id": {"$in": list(ann_user_ids)}}, {"_id": 0}).to_list(1000)
    user_display = {u['id']: (u.get('full_name') or u.get('email') or u['id']) for u in users}
    # Helper: format tags for a sentence, include user display, exclude skipped
    def format_sentence_tags(sent_id: str) -> str:
        arr = []
        for a in by_sid.get(sent_id, []):
            if a.get('skipped'):
                continue
            who = user_display.get(a.get('user_id',''), a.get('user_id',''))
            tags = a.get('tags', [])
            if isinstance(tags, list):
                for t in tags:
                    if isinstance(t, dict):
                        domain = (t.get('domain') or '').strip()
                        category = (t.get('category') or '').strip()
                        tag = (t.get('tag') or '').strip()
                        val = (t.get('valence') or '').strip().lower()
                    else:
                        domain = (getattr(t, 'domain', '') or '').strip()
                        category = (getattr(t, 'category', '') or '').strip()
                        tag = (getattr(t, 'tag', '') or '').strip()
                        val = (getattr(t, 'valence', '') or '').strip().lower()
                    if not domain or not tag:
                        continue
                    sign = '+' if val == 'positive' else '-'
                    arr.append(f"{domain}:{category}:{tag}({sign})@{who}")
        if not arr:
            return ''
        return f" [Tags: {', '.join(arr)}]"
    # Group sentences back to paragraphs per row_index
    para_map: Dict[int, Dict[str, Any]] = {}
    for s in sents:
        ri = int(s.get('row_index') or 0)
        if ri not in para_map:
            para_map[ri] = {'subject_id': s.get('subject_id', ''), 'parts': []}
        para_map[ri]['parts'].append(s.get('text','') + format_sentence_tags(s['id']))
    # Build CSV
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["row_index","subject_id","annotated_paragraph_text"]) 
    for ri in sorted(para_map.keys()):
        subj = para_map[ri]['subject_id']
        paragraph = ' '.join(para_map[ri]['parts']).strip()
        w.writerow([ri, subj, paragraph])
    out.seek(0)
    return StreamingResponse(io.BytesIO(out.getvalue().encode()), media_type='text/csv', headers={"Content-Disposition": f"attachment; filename=annotated_paragraphs_{doc['filename']}"})

# ========================
# Per-User Download Endpoints (Annotators)
# ========================
@api_router.get("/download/my-annotations-csv/{document_id}")
async def download_my_annotations_csv(document_id: str, current_user: User = Depends(get_current_user)):
    """Download annotator's own annotations for a document in inline CSV format"""
    document = await db.documents.find_one({"id": document_id})
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Get all sentences for this document
    sentences = await db.sentences.find({"document_id": document_id}, {"_id": 0}).sort([("row_index",1),("sentence_index",1)]).to_list(100000)
    sentence_map = {s["id"]: s for s in sentences}
    
    output = io.StringIO()
    writer = csv.writer(output)
    # Added timestamp column for annotation creation time; confidence is now per-tag
    writer.writerow(["document_id","sentence_id","subject_id","row_index","sentence_index","sentence_text","tag_domain","tag_category","tag","valence","confidence","notes","is_skipped","timestamp","duration_ms"]) 
    
    # For each sentence, get only current user's annotations
    for sid, s in sentence_map.items():
        anns = await db.annotations.find({"sentence_id": sid, "user_id": current_user.id}, {"_id": 0}).to_list(1000)
        if not anns:
            # No annotation by this user for this sentence - empty row
            writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), "","","","","","", False,"",""])
            continue
        
        for a in anns:
            duration_ms = a.get("duration_ms", "")
            timestamp = a.get("created_at", "")
            if a.get("skipped"):
                # Skipped annotation - no tags, mark as skipped=TRUE
                writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), "","","","","", a.get("notes",""), True, timestamp, duration_ms])
                continue
            
            tags = a.get("tags", [])
            if isinstance(tags, list) and tags:
                for t in tags:
                    domain = t.get("domain") if isinstance(t, dict) else getattr(t, "domain", "")
                    category = t.get("category") if isinstance(t, dict) else getattr(t, "category", "")
                    tag = t.get("tag") if isinstance(t, dict) else getattr(t, "tag", "")
                    valence = t.get("valence") if isinstance(t, dict) else getattr(t, "valence", "")
                    # Per-tag confidence (1-5 scale)
                    tag_confidence = t.get("confidence") if isinstance(t, dict) else getattr(t, "confidence", "")
                    writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), domain, category, tag, valence, tag_confidence, a.get("notes",""), False, timestamp, duration_ms])
            else:
                writer.writerow([document_id, sid, s.get("subject_id",""), s.get("row_index",""), s.get("sentence_index",""), s.get("text",""), "","","","","", a.get("notes",""), False, timestamp, duration_ms])
    
    output.seek(0)
    filename = f"my_annotations_{current_user.email.split('@')[0]}_{document.get('filename', 'export')}.csv"
    return StreamingResponse(io.BytesIO(output.getvalue().encode()), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.get("/download/my-annotated-paragraphs/{document_id}")
async def download_my_annotated_paragraphs(document_id: str, current_user: User = Depends(get_current_user)):
    """Download annotator's own annotations reconstructed as paragraphs"""
    doc = await db.documents.find_one({"id": document_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Fetch sentences ordered by original row/sentence order
    sents = await db.sentences.find({"document_id": document_id}, {"_id": 0}).sort([("row_index",1),("sentence_index",1)]).to_list(100000)
    if not sents:
        raise HTTPException(status_code=404, detail="No sentences for document")
    
    ids = [s['id'] for s in sents]
    # Fetch only current user's annotations
    anns = await db.annotations.find({"sentence_id": {"$in": ids}, "user_id": current_user.id}, {"_id": 0}).to_list(200000)
    
    by_sid: Dict[str, List[Dict[str, Any]]] = {}
    for a in anns:
        by_sid.setdefault(a['sentence_id'], []).append(a)
    
    # User display name
    user_display = current_user.full_name or current_user.email or current_user.id
    
    # Helper: format tags for a sentence (include skipped marker, timestamp, and per-tag confidence)
    def format_sentence_tags(sent_id: str) -> str:
        arr = []
        skipped_timestamps = []
        for a in by_sid.get(sent_id, []):
            timestamp = a.get('created_at', '')
            if a.get('skipped'):
                # Record skipped annotation with timestamp
                skipped_timestamps.append(timestamp)
                continue
            tags = a.get('tags', [])
            if isinstance(tags, list):
                for t in tags:
                    if isinstance(t, dict):
                        domain = (t.get('domain') or '').strip()
                        category = (t.get('category') or '').strip()
                        tag = (t.get('tag') or '').strip()
                        val = (t.get('valence') or '').strip().lower()
                        conf = t.get('confidence', '')
                    else:
                        domain = (getattr(t, 'domain', '') or '').strip()
                        category = (getattr(t, 'category', '') or '').strip()
                        tag = (getattr(t, 'tag', '') or '').strip()
                        val = (getattr(t, 'valence', '') or '').strip().lower()
                        conf = getattr(t, 'confidence', '')
                    if not domain or not tag:
                        continue
                    sign = '+' if val == 'positive' else '-'
                    # Include per-tag confidence in format: Domain:Category:Tag(+/-,conf=X)@User@Timestamp
                    conf_str = f",conf={conf}" if conf else ""
                    arr.append(f"{domain}:{category}:{tag}({sign}{conf_str})@{user_display}@{timestamp}")
        
        result_parts = []
        if arr:
            result_parts.append(f"[Tags: {', '.join(arr)}]")
        if skipped_timestamps:
            # Show skipped annotations with their timestamps
            for ts in skipped_timestamps:
                result_parts.append(f"[SKIPPED@{user_display}@{ts}]")
        
        if not result_parts:
            return ''
        return ' ' + ' '.join(result_parts)
    
    # Group sentences back to paragraphs per row_index
    para_map: Dict[int, Dict[str, Any]] = {}
    for s in sents:
        ri = int(s.get('row_index') or 0)
        if ri not in para_map:
            para_map[ri] = {'subject_id': s.get('subject_id', ''), 'parts': []}
        para_map[ri]['parts'].append(s.get('text','') + format_sentence_tags(s['id']))
    
    # Build CSV
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["row_index","subject_id","annotated_paragraph_text"]) 
    for ri in sorted(para_map.keys()):
        subj = para_map[ri]['subject_id']
        paragraph = ' '.join(para_map[ri]['parts']).strip()
        w.writerow([ri, subj, paragraph])
    
    out.seek(0)
    filename = f"my_paragraphs_{current_user.email.split('@')[0]}_{doc.get('filename', 'export')}"
    return StreamingResponse(io.BytesIO(out.getvalue().encode()), media_type='text/csv', headers={"Content-Disposition": f"attachment; filename={filename}"})


# ========================
# Analytics (partial)
# ========================
@api_router.get("/admin/analytics/users")
async def get_user_analytics(current_user: User = Depends(get_admin_user)):
    """Get per-user analytics (admin only)"""
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    analytics = {}
    
    for user in users:
        user_annotations = await db.annotations.find({"user_id": user["id"]}, {"_id": 0}).to_list(10000)
        total_annotations = len(user_annotations)
        tagged_annotations = len([a for a in user_annotations if not a.get("skipped", False)])
        skipped_annotations = len([a for a in user_annotations if a.get("skipped", False)])
        
        analytics[user["id"]] = {
            "user": {
                "id": user["id"],
                "email": user["email"],
                "full_name": user.get("full_name", ""),
                "role": user.get("role", "annotator")
            },
            "total_annotations": total_annotations,
            "tagged_annotations": tagged_annotations,
            "skipped_annotations": skipped_annotations
        }
    
    return analytics

@api_router.get("/analytics/overview")
async def analytics_overview(current_user: User = Depends(get_current_user)):
    total_documents = await db.documents.count_documents({})
    total_sentences = await db.sentences.count_documents({})
    total_annotations = await db.annotations.count_documents({})
    tagged = await db.annotations.count_documents({"skipped": False})
    skipped = await db.annotations.count_documents({"skipped": True})
    annotators = len(await db.annotations.distinct("user_id"))
    return {
        "total_documents": total_documents,
        "total_sentences": total_sentences,
        "total_annotations": total_annotations,
        "tagged": tagged,
        "skipped": skipped,
        "annotators": annotators
    }

@api_router.get("/analytics/enhanced")
async def get_enhanced_analytics(current_user: User = Depends(get_current_user)):
    users = await db.users.find({}, {"_id": 0}).to_list(1000)
    per_user = []
    for user in users:
        user_annotations = await db.annotations.count_documents({"user_id": user["id"]})
        tagged = await db.annotations.count_documents({"user_id": user["id"], "skipped": False})
        skipped = await db.annotations.count_documents({"user_id": user["id"], "skipped": True})
        per_user.append({
            "user_id": user["id"],
            "full_name": user.get("full_name", user.get("email", "")),
            "total": user_annotations,
            "tagged": tagged,
            "skipped": skipped,
            "sentences_left": 0
        })
    total_sentences = await db.sentences.count_documents({})
    annotated_sentences = len(await db.annotations.distinct("sentence_id"))
    sentences_left_overall = total_sentences - annotated_sentences
    irr_pairs = []
    return {"per_user": per_user, "sentences_left_overall": sentences_left_overall, "irr_pairs": irr_pairs}

@api_router.get("/analytics/tag-prevalence-chart")
async def get_tag_prevalence_chart(current_user: Optional[User] = Depends(get_current_user_optional), token: Optional[str] = None):
    # Allow ?token for image rendering in <img> tags
    if token and not current_user:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                u = await db.users.find_one({"id": user_id}, {"_id": 0})
                if u:
                    current_user = User(**u)
                else:
                    raise HTTPException(status_code=401, detail="Invalid token user")
        except Exception:
            raise HTTPException(status_code=401, detail="Invalid token")
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Get real tag counts from annotations
    annotations = await db.annotations.find({"skipped": {"$ne": True}}, {"_id": 0, "tags": 1}).to_list(100000)
    
    domain_counts = {}
    for ann in annotations:
        tags = ann.get('tags', [])
        if isinstance(tags, list):
            for t in tags:
                domain = t.get('domain') if isinstance(t, dict) else getattr(t, 'domain', '')
                if domain:
                    domain_counts[domain] = domain_counts.get(domain, 0) + 1
    
    # Use SDOH domains in order
    domain_labels = list(SDOH_DOMAINS)
    domain_values = [domain_counts.get(d, 0) for d in domain_labels]
    
    # Shorten labels for display
    short_labels = [d.split()[0] if len(d) > 15 else d for d in domain_labels]
    
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from io import BytesIO
    
    colors = ['#2563eb', '#9333ea', '#059669', '#dc2626', '#f59e0b']
    
    plt.figure(figsize=(12, 6))
    bars = plt.bar(short_labels, domain_values, color=colors[:len(domain_labels)])
    plt.title('Tag Count by SDOH Domain', fontsize=14, fontweight='bold')
    plt.ylabel('Number of Tags')
    plt.xticks(rotation=15, ha='right')
    
    # Add value labels on bars
    for bar, val in zip(bars, domain_values):
        if val > 0:
            plt.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, str(val), 
                    ha='center', va='bottom', fontsize=10)
    
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    plt.close()
    buf.seek(0)
    return StreamingResponse(buf, media_type='image/png')

@api_router.get("/analytics/domain-tag-stats")
async def get_domain_tag_stats(current_user: User = Depends(get_current_user)):
    """Get detailed tag counts per domain and per specific tag"""
    annotations = await db.annotations.find({"skipped": {"$ne": True}}, {"_id": 0, "tags": 1}).to_list(100000)
    
    # Count tags per domain and per specific tag
    domain_counts = {}
    tag_counts = {}  # {domain: {category: {tag: count}}}
    
    for ann in annotations:
        tags = ann.get('tags', [])
        if isinstance(tags, list):
            for t in tags:
                if isinstance(t, dict):
                    domain = t.get('domain', '')
                    category = t.get('category', '')
                    tag = t.get('tag', '')
                else:
                    domain = getattr(t, 'domain', '')
                    category = getattr(t, 'category', '')
                    tag = getattr(t, 'tag', '')
                
                if domain:
                    domain_counts[domain] = domain_counts.get(domain, 0) + 1
                    if domain not in tag_counts:
                        tag_counts[domain] = {}
                    if category not in tag_counts[domain]:
                        tag_counts[domain][category] = {}
                    if tag:
                        tag_counts[domain][category][tag] = tag_counts[domain][category].get(tag, 0) + 1
    
    # Format response
    result = {
        "domain_totals": domain_counts,
        "tag_details": tag_counts,
        "domains": list(SDOH_DOMAINS)
    }
    return result

@api_router.get("/analytics/domain-chart/{domain_name}")
async def get_domain_chart(domain_name: str, current_user: Optional[User] = Depends(get_current_user_optional), token: Optional[str] = None):
    """Get chart showing tag counts for a specific domain"""
    if token and not current_user:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                u = await db.users.find_one({"id": user_id}, {"_id": 0})
                if u:
                    current_user = User(**u)
        except Exception:
            pass
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Get annotations with this domain
    annotations = await db.annotations.find({"skipped": {"$ne": True}}, {"_id": 0, "tags": 1}).to_list(100000)
    
    tag_counts = {}
    for ann in annotations:
        tags = ann.get('tags', [])
        if isinstance(tags, list):
            for t in tags:
                if isinstance(t, dict):
                    domain = t.get('domain', '')
                    tag = t.get('tag', '')
                else:
                    domain = getattr(t, 'domain', '')
                    tag = getattr(t, 'tag', '')
                
                if domain == domain_name and tag:
                    tag_counts[tag] = tag_counts.get(tag, 0) + 1
    
    if not tag_counts:
        # Return empty chart
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from io import BytesIO
        plt.figure(figsize=(10, 4))
        plt.text(0.5, 0.5, 'No data available', ha='center', va='center', fontsize=14)
        plt.axis('off')
        buf = BytesIO()
        plt.savefig(buf, format='png', dpi=100)
        plt.close()
        buf.seek(0)
        return StreamingResponse(buf, media_type='image/png')
    
    # Sort by count and take top 15
    sorted_tags = sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:15]
    labels = [t[0][:25] + '...' if len(t[0]) > 25 else t[0] for t in sorted_tags]
    values = [t[1] for t in sorted_tags]
    
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from io import BytesIO
    
    plt.figure(figsize=(12, max(4, len(labels) * 0.4)))
    bars = plt.barh(range(len(labels)), values, color='#3b82f6')
    plt.yticks(range(len(labels)), labels)
    plt.xlabel('Count')
    plt.title(f'Tag Distribution: {domain_name}', fontsize=12, fontweight='bold')
    
    # Add value labels
    for i, (bar, val) in enumerate(zip(bars, values)):
        plt.text(val + 0.3, i, str(val), va='center', fontsize=9)
    
    plt.gca().invert_yaxis()
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    plt.close()
    buf.seek(0)
    return StreamingResponse(buf, media_type='image/png')

@api_router.get("/analytics/document-user-progress/{document_id}")
async def get_document_user_progress(document_id: str, current_user: User = Depends(get_current_user)):
    """Get annotation progress per user for a document"""
    doc = await db.documents.find_one({"id": document_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Get total sentences
    total_sentences = await db.sentences.count_documents({"document_id": document_id})
    if total_sentences == 0:
        return {"document_id": document_id, "filename": doc.get('filename', ''), "total_sentences": 0, "user_progress": []}
    
    # Get all sentence IDs
    sentence_ids = await db.sentences.distinct("id", {"document_id": document_id})
    
    # Get assigned users
    assigned_users = doc.get('assigned_users', [])
    
    # Get all users who have annotated this document
    annotator_ids = await db.annotations.distinct("user_id", {"sentence_id": {"$in": sentence_ids}})
    
    # Combine assigned and actual annotators
    all_user_ids = list(set(assigned_users + annotator_ids))
    
    # Get user info
    users = await db.users.find({"id": {"$in": all_user_ids}}, {"_id": 0}).to_list(100)
    user_map = {u['id']: u.get('full_name') or u.get('email') or u['id'] for u in users}
    
    # Calculate progress per user
    user_progress = []
    for user_id in all_user_ids:
        annotated_count = len(await db.annotations.distinct("sentence_id", {
            "sentence_id": {"$in": sentence_ids},
            "user_id": user_id
        }))
        user_progress.append({
            "user_id": user_id,
            "user_name": user_map.get(user_id, user_id),
            "annotated": annotated_count,
            "total": total_sentences,
            "progress": round(annotated_count / total_sentences * 100, 1)
        })
    
    # Sort by progress descending
    user_progress.sort(key=lambda x: x['progress'], reverse=True)
    
    return {
        "document_id": document_id,
        "filename": doc.get('filename', ''),
        "total_sentences": total_sentences,
        "user_progress": user_progress
    }

@api_router.get("/analytics/all-documents-user-progress")
async def get_all_documents_user_progress(current_user: User = Depends(get_current_user)):
    """Get annotation progress per user for all documents"""
    docs = await db.documents.find({}, {"_id": 0}).to_list(1000)
    
    results = []
    for doc in docs:
        document_id = doc['id']
        total_sentences = await db.sentences.count_documents({"document_id": document_id})
        if total_sentences == 0:
            continue
        
        sentence_ids = await db.sentences.distinct("id", {"document_id": document_id})
        assigned_users = doc.get('assigned_users', [])
        annotator_ids = await db.annotations.distinct("user_id", {"sentence_id": {"$in": sentence_ids}})
        all_user_ids = list(set(assigned_users + annotator_ids))
        
        if not all_user_ids:
            continue
        
        users = await db.users.find({"id": {"$in": all_user_ids}}, {"_id": 0}).to_list(100)
        user_map = {u['id']: u.get('full_name') or u.get('email') or u['id'] for u in users}
        
        user_progress = []
        for user_id in all_user_ids:
            annotated_count = len(await db.annotations.distinct("sentence_id", {
                "sentence_id": {"$in": sentence_ids},
                "user_id": user_id
            }))
            user_progress.append({
                "user_id": user_id,
                "user_name": user_map.get(user_id, user_id),
                "annotated": annotated_count,
                "total": total_sentences,
                "progress": round(annotated_count / total_sentences * 100, 1)
            })
        
        user_progress.sort(key=lambda x: x['progress'], reverse=True)
        
        results.append({
            "document_id": document_id,
            "filename": doc.get('filename', ''),
            "total_sentences": total_sentences,
            "user_progress": user_progress
        })
    
    return results

@api_router.get("/analytics/projects")
async def get_projects_analytics(current_user: User = Depends(get_current_user)):
    docs = await db.documents.find({}, {"_id": 0}).to_list(10000)
    projects: Dict[str, Dict[str, Any]] = {}
    for d in docs:
        pname = d.get('project_name') or 'Unassigned'
        if pname not in projects:
            projects[pname] = {
                'project_name': pname,
                'documents': [],
                'document_ids': [],
                'documents_count': 0,
                'total_sentences': 0,
                'annotated_sentences': 0,
                'annotators_count': 0,
                'last_activity': None,
            }
        projects[pname]['documents'].append({'id': d['id'], 'filename': d.get('filename')})
        projects[pname]['document_ids'].append(d['id'])
    if not projects:
        return []
    for pname, p in projects.items():
        ids = p['document_ids']
        p['documents_count'] = len(ids)
        p['total_sentences'] = await db.sentences.count_documents({"document_id": {"$in": ids}})
        if p['total_sentences'] == 0:
            continue
        sentence_ids = await db.sentences.distinct("id", {"document_id": {"$in": ids}})
        if sentence_ids:
            annotated_sentence_ids = await db.annotations.distinct("sentence_id", {"sentence_id": {"$in": sentence_ids}})
            p['annotated_sentences'] = len(annotated_sentence_ids)
            annotators = await db.annotations.distinct("user_id", {"sentence_id": {"$in": sentence_ids}})
            p['annotators_count'] = len(annotators)
            last = await db.annotations.find({"sentence_id": {"$in": sentence_ids}}, {"_id": 0, "created_at": 1}).sort("created_at", -1).limit(1).to_list(1)
            if last:
                p['last_activity'] = str(last[0].get('created_at'))
    result = []
    for pname, p in projects.items():
        progress = 0.0 if p['total_sentences'] == 0 else p['annotated_sentences'] / p['total_sentences']
        result.append({
            'project_name': pname,
            'documents_count': p['documents_count'],
            'total_sentences': p['total_sentences'],
            'annotated_sentences': p['annotated_sentences'],
            'progress': progress,
            'annotators_count': p['annotators_count'],
            'last_activity': p['last_activity'],
        })
    result.sort(key=lambda x: x['annotated_sentences'], reverse=True)
    return result

@api_router.get("/analytics/projects-chart")
async def get_projects_chart(current_user: Optional[User] = Depends(get_current_user_optional), token: Optional[str] = None):
    """
    Stacked bar chart (Option B): Completed vs Remaining sentences per project.
    Supports token query param for environments where <img> requests cannot send Authorization headers.
    """
    if token and not current_user:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                u = await db.users.find_one({"id": user_id}, {"_id": 0})
                if u:
                    if isinstance(u.get('created_at'), datetime):
                        u['created_at'] = u['created_at'].isoformat()
                    current_user = User(**u)
                else:
                    raise HTTPException(status_code=401, detail="Invalid token user")
        except Exception:
            raise HTTPException(status_code=401, detail="Invalid token")
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from io import BytesIO
    data = await get_projects_analytics(current_user)
    if not data:
        data = []
    data = data[:20]
    labels = [d['project_name'] for d in data]
    completed = [int(d.get('annotated_sentences') or 0) for d in data]
    totals = [int(d.get('total_sentences') or 0) for d in data]
    remaining = [max(t - c, 0) for t, c in zip(totals, completed)]
    plt.figure(figsize=(12, 7))
    y_pos = range(len(labels))
    plt.barh(y_pos, completed, color='#16a34a', label='Completed (Annotated)')
    plt.barh(y_pos, remaining, left=completed, color='#93c5fd', label='Remaining')
    plt.yticks(y_pos, labels)
    plt.xlabel('Sentences')
    plt.title('Project Progress: Completed vs Remaining Sentences')
    plt.legend(loc='lower right')
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    return StreamingResponse(buf, media_type='image/png')

# Include API router with /api prefix for all endpoints

@api_router.post("/admin/resources/upload")
async def upload_resource(file: UploadFile = File(...), current_user: User = Depends(get_admin_user)):
    if not file or not getattr(file, 'filename', None):
        raise HTTPException(status_code=400, detail="No file provided")
    ext = (file.filename.split('.')[-1] or '').lower()
    if ext not in ALLOWED_RESOURCE_EXT:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: .{ext}")
    data = await file.read()
    
    # Upload to GridFS
    file_id = await fs_bucket.upload_from_stream(
        file.filename,
        data,
        metadata={
            'uploaded_by': current_user.id,
            'content_type': file.content_type or 'application/octet-stream',
            'size': len(data),
            'kind': 'file'
        }
    )
    
    rid = str(file_id)
    await db.resources_meta.insert_one({
        'id': rid,
        'filename': file.filename,
        'content_type': file.content_type or 'application/octet-stream',
        'uploaded_by': current_user.id,
        'uploaded_at': datetime.utcnow().isoformat(),
        'size': len(data),
        'kind': 'file'
    })
    return {'id': rid}

# Optional: External resource links (Google Docs/Slides/Sheets)
class ExternalResource(BaseModel):
    title: str
    url: str

@api_router.post("/admin/resources/link")
async def add_resource_link(payload: ExternalResource, current_user: User = Depends(get_admin_user)):
    try:
        u = urlparse(payload.url)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid URL")
    if not u.scheme.startswith('http'):
        raise HTTPException(status_code=400, detail="Invalid URL scheme")
    rid = str(uuid.uuid4())
    await db.resources_meta.insert_one({
        'id': rid,
        'filename': payload.title,
        'content_type': 'text/uri-list',
        'uploaded_by': current_user.id,
        'uploaded_at': datetime.utcnow().isoformat(),
        'size': 0,
        'link_url': payload.url,
        'kind': 'link'
    })
    return {'id': rid}

@api_router.get("/resources")
async def list_resources(
    q: Optional[str] = None,
    kind: Optional[str] = None, # 'file' or 'link'
    mime: Optional[str] = None, # 'image', 'pdf', 'office'
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_user)
):
    query: Dict[str, Any] = {}
    if q:
        query['filename'] = {"$regex": q, "$options": "i"}
    if kind:
        if kind == 'file':
            query['kind'] = {"$ne": "link"}
        elif kind == 'link':
            query['kind'] = 'link'
    if mime:
        if mime == 'image':
            query['content_type'] = {"$regex": r"^image/"}
        elif mime == 'pdf':
            query['content_type'] = 'application/pdf'
        elif mime == 'office':
            query['content_type'] = {"$in": [
                'application/msword','application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/vnd.ms-excel','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-powerpoint','application/vnd.openxmlformats-officedocument.presentationml.presentation'
            ]}
    skip = max(0, (page-1) * page_size)
    total = await db.resources_meta.count_documents(query)
    cursor = db.resources_meta.find(query, {"_id": 0}).sort("uploaded_at", -1).skip(skip).limit(page_size)
    items = await cursor.to_list(page_size)
    # ensure kind for legacy entries
    for it in items:
        if 'kind' not in it:
            it['kind'] = 'file'
    return {"items": items, "total": total, "page": page, "page_size": page_size}

@api_router.get("/resources/{resource_id}")
async def get_resource_meta(resource_id: str, current_user: User = Depends(get_current_user)):
    meta = await db.resources_meta.find_one({"id": resource_id}, {"_id": 0})
    if not meta:
        raise HTTPException(status_code=404, detail="Not found")
    return meta

@api_router.get("/resources/{resource_id}/download")
async def download_resource(resource_id: str, current_user: Optional[User] = Depends(get_current_user_optional), token: Optional[str] = None):
    # Allow ?token for image/pdf previews in <img>/<iframe>
    if token and not current_user:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                u = await db.users.find_one({"id": user_id}, {"_id": 0})
                if not u:
                    raise HTTPException(status_code=401, detail="Invalid token user")
                current_user = User(**u)
        except Exception:
            raise HTTPException(status_code=401, detail="Invalid token")
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        oid = ObjectId(resource_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Not found")
    meta = await db.resources_meta.find_one({"id": resource_id}, {"_id": 0})
    if not meta:
        raise HTTPException(status_code=404, detail="Not found")
    
    # Download from GridFS
    buffer = io.BytesIO()
    await fs_bucket.download_to_stream(oid, buffer)
    buffer.seek(0)
    
    return StreamingResponse(buffer, media_type=meta.get('content_type') or 'application/octet-stream', headers={
        "Content-Disposition": f"inline; filename={meta['filename']}"
    })


@api_router.delete("/admin/resources/{resource_id}")
async def delete_resource(resource_id: str, current_user: User = Depends(get_admin_user)):
    """Delete a resource (admin only)"""
    # First check if resource exists and get its type
    meta = await db.resources_meta.find_one({"id": resource_id}, {"_id": 0})
    if not meta:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    # Only try to delete from GridFS if it's a file (not a link)
    if meta.get('kind') != 'link':
        try:
            oid = ObjectId(resource_id)
            await fs_bucket.delete(oid)
        except Exception:
            # If not found in GridFS or invalid ObjectId, continue to delete metadata
            pass
    
    # Delete metadata
    result = await db.resources_meta.delete_one({"id": resource_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    return {"message": "Resource deleted successfully"}

@api_router.get("/resources/{resource_id}/preview")
async def preview_resource(resource_id: str, current_user: Optional[User] = Depends(get_current_user_optional), token: Optional[str] = None):
    """Get HTML preview of a Word document or Excel file"""
    # Allow ?token for iframe previews
    if token and not current_user:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                u = await db.users.find_one({"id": user_id}, {"_id": 0})
                if u:
                    current_user = User(**u)
        except Exception:
            pass
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        oid = ObjectId(resource_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid resource ID")
    
    meta = await db.resources_meta.find_one({"id": resource_id}, {"_id": 0})
    if not meta:
        raise HTTPException(status_code=404, detail="Not found")
    
    content_type = meta.get('content_type', '')
    filename = meta.get('filename', '').lower()
    
    # Check if it's a Word document
    is_word = 'word' in content_type.lower() or filename.endswith(('.doc', '.docx'))
    # Check if it's an Excel file
    is_excel = 'spreadsheet' in content_type.lower() or 'excel' in content_type.lower() or filename.endswith(('.xls', '.xlsx'))
    
    if not is_word and not is_excel:
        raise HTTPException(status_code=400, detail="Preview only available for Word documents and Excel files")
    
    # Download from GridFS
    buffer = io.BytesIO()
    await fs_bucket.download_to_stream(oid, buffer)
    buffer.seek(0)
    
    if is_word:
        # Convert Word to HTML using mammoth
        try:
            result = mammoth.convert_to_html(buffer)
            html_content = result.value
            
            full_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; padding: 20px; max-width: 800px; margin: 0 auto; }}
                    p {{ margin: 10px 0; }}
                    h1, h2, h3 {{ margin: 15px 0 10px 0; }}
                    table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
                    td, th {{ border: 1px solid #ddd; padding: 8px; }}
                    th {{ background-color: #f2f2f2; }}
                </style>
            </head>
            <body>
                {html_content}
            </body>
            </html>
            """
            return HTMLResponse(content=full_html)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error converting document: {str(e)}")
    
    elif is_excel:
        # Convert Excel to HTML table (first 10 rows)
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.load_workbook(buffer, read_only=True, data_only=True)
            sheet = wb.active
            
            rows_html = []
            row_count = 0
            max_rows = 10
            
            for row in sheet.iter_rows(max_row=max_rows + 1):
                cells = []
                for cell in row:
                    value = cell.value if cell.value is not None else ''
                    # Escape HTML
                    value = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    if row_count == 0:
                        cells.append(f'<th>{value}</th>')
                    else:
                        cells.append(f'<td>{value}</td>')
                rows_html.append(f'<tr>{"".join(cells)}</tr>')
                row_count += 1
                if row_count > max_rows:
                    break
            
            total_rows = sheet.max_row or 0
            total_cols = sheet.max_column or 0
            
            wb.close()
            
            full_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; padding: 20px; }}
                    .info {{ color: #666; margin-bottom: 15px; font-size: 14px; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    td, th {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f8f9fa; font-weight: bold; }}
                    tr:nth-child(even) {{ background-color: #f9f9f9; }}
                    .note {{ color: #888; font-style: italic; margin-top: 10px; font-size: 12px; }}
                </style>
            </head>
            <body>
                <div class="info">Excel Preview: {total_rows} rows × {total_cols} columns</div>
                <table>
                    {"".join(rows_html)}
                </table>
                {"<p class='note'>Showing first 10 rows. Download file to see all data.</p>" if total_rows > 10 else ""}
            </body>
            </html>
            """
            return HTMLResponse(content=full_html)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error reading Excel file: {str(e)}")


app.include_router(api_router)

@api_router.get("/admin/settings/default-project")
async def get_default_project(current_user: User = Depends(get_admin_user)):
    value = await get_default_project_name()
    return {"value": value}

@api_router.put("/admin/settings/default-project")
async def set_default_project(name: str = Query(..., min_length=1, max_length=200), current_user: User = Depends(get_admin_user)):
    await db.settings.update_one({"key": "default_project_name"}, {"$set": {"key": "default_project_name", "value": name, "updated_at": datetime.utcnow().isoformat()}}, upsert=True)
    return {"value": name}

@api_router.post("/admin/documents/reassign-to-default")
async def reassign_documents_to_default(current_user: User = Depends(get_admin_user)):
    default_name = await get_default_project_name()
    await db.documents.update_many({}, {"$set": {"project_name": default_name}})
    return {"updated": True, "project_name": default_name}
